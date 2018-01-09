# -*- coding: utf-8 -*-
import os
from collections import OrderedDict


from django.conf import settings

from openpyxl import load_workbook
from openpyxl.styles import Font, NamedStyle, PatternFill
from openpyxl.utils import column_index_from_string, coordinate_from_string, get_column_letter

from ..models import DynamicRegistry
from ..utils import insert_filename_pre_extension, now_for_filename, xlsx_download_response_factory


# Value allowing to skip a cell while continuing to read config of other columns.
DATA_SLUG_EMPTY_CELL = "#"


class AbstractExportExcel(object):
    """Abstract base class for Excel exports."""

    DEFAULT_FILENAME = "easynut-export.xlsx"  # Default name of the file.

    # Minimum number of cols to apply formatting to.
    APPLY_STYLE_MIN_NUM_COLS = 30

    # Available named styles.
    STYLE_HEADING = "heading"  # For heading rows.
    STYLE_VERBOSE = "advanced"  # For advanced information (cf. ``VERBOSE``).
    STYLE_MSF_ID = "msf_id"  # For the special field ``MSF ID``.
    STYLE_DATE = "date"  # For the special field ``Date``.

    # Named styles config.
    HEADING_BG_COLOR = "DDDDFF"
    VERBOSE_FG_COLOR = "999999"
    MSF_ID_BG_COLOR = "FFDDDD"
    DATE_BG_COLOR = "DDFFDD"

    def __init__(self):
        self.book = None  # The Excel workbook.
        self.filename = None  # The name of the file.

        self.styles = {}  # Named styles for formatting Excel cells.

        # Initialize the file name.
        self.init_filename()

    def apply_style_to_cols(self, style, sheet, min_col, max_col, min_row=None, max_row=None):
        """Apply a style to columns."""
        for row_cells in sheet.iter_rows(min_col=min_col, max_col=max_col, min_row=min_row, max_row=max_row):
            for cell in row_cells:
                cell.style = style

    def apply_style_to_rows(self, style, sheet, min_row, max_row, min_col=None, max_col=None):
        self.apply_style_to_cols(style, sheet, min_col, max_col, min_row=min_row, max_row=max_row)
        # for col_cells in sheet.iter_cols(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        #     for cell in col_cells:
        #         cell.style = style

    def cell_name_to_col_row(self, cell):
        """Convert a cell name into ``(col, row)`` coordinate. E.g. ``B5`` => ``(2, 5)``."""
        col, row = coordinate_from_string(cell)
        col = column_index_from_string(col)
        return col, row

    def cell_col_row_to_name(self, col, row):
        """Convert ``(col, row)`` coordinate into a cell name. E.g. ``(2, 5)`` => ``B5``."""
        col_letter = get_column_letter(col)
        return "{}{}".format(col_letter, row)

    def create_sheet(self, title):
        """Create a new sheet in the current workbook."""
        return self.book.create_sheet(title=title)

    def get_sheet(self, index):
        """Return a sheet identified by its index."""
        return self.book.worksheets[index]

    def get_sheet_names(self):
        """Return the list of sheet names."""
        return self.book.sheetnames

    def get_style(self, name):
        """Return a named style to format an Excel cell."""
        if name not in self.styles:
            self.styles[name] = NamedStyle(name=name)
            if name == self.STYLE_HEADING:
                self.styles[name].font = Font(bold=True)
                self.styles[name].fill = PatternFill(
                    fill_type="solid", start_color=self.HEADING_BG_COLOR, end_color=self.HEADING_BG_COLOR
                )
            elif name == self.STYLE_VERBOSE:
                self.styles[name].font = Font(color=self.VERBOSE_FG_COLOR)
            elif name == self.STYLE_MSF_ID:
                self.styles[name].fill = PatternFill(
                    fill_type="solid", start_color=self.MSF_ID_BG_COLOR, end_color=self.MSF_ID_BG_COLOR
                )
            elif name == self.STYLE_DATE:
                self.styles[name].fill = PatternFill(
                    fill_type="solid", start_color=self.DATE_BG_COLOR, end_color=self.DATE_BG_COLOR
                )
            else:
                return None

        return self.styles[name]

    def init_filename(self):
        """Initialize file name to its default value with "now"."""
        self.filename = insert_filename_pre_extension(self.DEFAULT_FILENAME, now_for_filename())

    def save(self, filename=None):
        """Save the Excel file (under ``settings.EXPORTS_ROOT``)."""
        self._save_common()
        if not filename:
            filename = self.filename

        # Save the file on the file system.
        file_path = os.path.join(settings.EXPORTS_ROOT, filename)
        self.book.save(file_path)

        # Return the path where the file has been saved.
        return file_path

    def save_to_response(self):
        """Save the Excel file in an HTTP response."""
        self._save_common()

        # Create a download response for an Excel (xlsx) file.
        response = xlsx_download_response_factory(self.filename)
        self.book.save(response)
        return response

    def _save_common(self):
        """Provide common steps for "save" methods."""
        if self.book is None:
            raise Exception("No Excel Workbook to save.")

        # Set active sheet to first one.
        self.book.active = 0


class AbstractExportExcelTemplate(AbstractExportExcel):
    """Abstract base class for Excel exports based on a template."""

    DEFAULT_TEMPLATE = None  # Default Excel template to use.

    def __init__(self, template=None, filename=None):
        super(AbstractExportExcelTemplate, self).__init__()
        self.template = template
        self.filename = filename

        # Config of the loaded template.
        self._config = OrderedDict()

        # Load the template (if not specified, load default one).
        self.load_template(template)

    def load_template(self, template=None):
        """Create the workbook from a template file (located under ``settings.EXPORTS_TEMPLATES_DIR``)."""
        # Set the current template.
        if template is None:
            if self.DEFAULT_TEMPLATE is None:
                raise Exception("No template defined.")
            template = self.DEFAULT_TEMPLATE
        self.template = template

        # Load the template.
        file_path = os.path.join(settings.EXPORTS_TEMPLATES_DIR, self.template)
        self.book = load_workbook(file_path)

        # Read the template config.
        self._init_config()

    def populate(self):
        """Populate the template with data."""
        if self.book is None:
            raise Exception("No template loaded.")
        # Must be implemented in concrete classes. Don't raise ``NotImplemented`` so that they can override this.

    def _init_config(self):
        """
        Read the template config.

        - The config sheet must be the first sheet in the workbook.
        - It is automatically removed.
        - See the concrete class for more specs.
        """
        raise NotImplemented()

    def _sheets_iterator(self, for_config=False):
        """Iterate over configured sheets."""
        for index, config in self._config.iteritems():
            yield index, self.get_sheet(index), config

    def _update_models_fields(self, sheet_index, data_slug):
        """Register the models and fields for a given sheet."""
        # Split the data slug into ``(model_id, field_id)``.
        model_id, field_id = DynamicRegistry.split_data_slug(data_slug)

        # Register the model and field.
        config = self._config[sheet_index]["models_fields"]
        if model_id not in config:
            config[model_id] = []
        if field_id not in config[model_id]:
            config[model_id].append(field_id)