# -*- coding: utf-8 -*-
import json

from openpyxl import Workbook

from ..models import MSF_ID_VERBOSE_NAME, DATE_VERBOSE_NAME, DynamicRegistry

from .base import AbstractExportExcel


class ExportDataModel(AbstractExportExcel):
    """Create an Excel file containing a list of all models and fields with their data slug."""

    DEFAULT_FILENAME = "easynut-data-model.xlsx"  # Default name for the file.
    VERBOSE = True  # Whether to include advanced information.

    # Headings. Use ``None`` to skip a column.
    DATA_SLUGS_HEADINGS = ["Model Name", "Field Name", "Data Slug"]
    DATA_SLUGS_HEADINGS_VERBOSE = [None, "Model ID", "Field ID", "Position", "Type", "Values List"]
    MODELS_HEADINGS = ["Model Name"]
    MODELS_HEADINGS_VERBOSE = ["Model ID"]

    def generate(self):
        """Generate the export."""
        self.book = Workbook()
        DynamicRegistry.load_models_config()

        self._generate_data_slugs()
        self._generate_models()

        return self.book

    def _generate_data_slugs(self):
        """Generate the data slugs sheet."""
        sheet = self.get_sheet(0)
        sheet.title = "Data Slugs"

        headings = self.DATA_SLUGS_HEADINGS
        if self.VERBOSE:
            headings += self.DATA_SLUGS_HEADINGS_VERBOSE
        self._write_headings(sheet, headings)

        self._write_values_data_slugs(sheet)

    def _generate_models(self):
        """Generate the models sheet."""
        sheet = self.create_sheet("Models")

        headings = self.MODELS_HEADINGS
        if self.VERBOSE:
            headings += self.MODELS_HEADINGS_VERBOSE
        self._write_headings(sheet, headings)

        self._write_values_models(sheet)

    def _save_common(self):
        """Provide common steps for "save" methods."""
        if self.book is None:
            self.generate()
        super(ExportDataModel, self)._save_common()

    def _write_headings(self, sheet, headings):
        """Write heading row."""
        col = 0; row = 1  # NOQA
        for value in headings:
            col += 1
            if value is not None:  # None = skip column.
                sheet.cell(column=col, row=row).value = value

        # Apply heading style and freeze heading row.
        style = self.get_style(self.STYLE_HEADING)
        max_col = max(len(headings), self.APPLY_STYLE_MIN_NUM_COLS)  # Apply to at least a min number of columns.
        self.apply_style_to_rows(style, sheet, min_row=row, max_row=row, max_col=max_col)
        sheet.freeze_panes = sheet.cell(column=1, row=row + 1)

    def _write_values(self, sheet, row, values):
        """Write values on the given row."""
        col = 0
        for value in values:
            col += 1
            if value is not None:  # None = skip column.
                sheet.cell(column=col, row=row).value = value

    def _write_values_data_slugs(self, sheet):
        """Write data slugs values."""
        # Loop over all models and fields config.
        min_row = 2  # Skip heading row.
        row = min_row - 1    # To start the loops with the increment, easier to read.
        for model_id, model_config in DynamicRegistry.models_config.iteritems():
            for field_id, field_config in model_config.fields_config.iteritems():
                row += 1

                # Build values.
                values = [
                    model_config.name,
                    field_config.name,
                    field_config.data_slug,
                ]
                verbose_min_col = len(values) + 1
                if self.VERBOSE:
                    values += [
                        None,
                        model_id,
                        field_id,
                        field_config.position,
                        field_config.kind,
                        json.dumps(field_config.values_list) if field_config.values_list else "",
                    ]
                    verbose_max_col = len(values) + 1

                # Write values.
                self._write_values(sheet, row, values)

                # Apply special fields style.
                style = None
                if field_config.name == MSF_ID_VERBOSE_NAME:
                    style = self.get_style(self.STYLE_MSF_ID)
                elif field_config.name == DATE_VERBOSE_NAME:
                    style = self.get_style(self.STYLE_DATE)
                if style is not None:
                    self.apply_style_to_rows(
                        style, sheet, min_row=row, max_row=row, min_col=1, max_col=verbose_min_col - 1
                    )

                # Apply verbose style.
                style = self.get_style(self.STYLE_VERBOSE)
                self.apply_style_to_cols(
                    style, sheet, min_col=verbose_min_col, max_col=verbose_max_col, min_row=min_row
                )

    def _write_values_models(self, sheet):
        """Write models values."""
        # Loop over all models config.
        min_row = 2  # Skip heading row.
        row = min_row - 1    # To start the loops with the increment, easier to read.
        for model_id, model_config in DynamicRegistry.models_config.iteritems():
            row += 1

            # Build values.
            values = [model_config.name]
            verbose_min_col = len(values) + 1
            if self.VERBOSE:
                values += [model_id]
                verbose_max_col = len(values) + 1

            # Write values.
            self._write_values(sheet, row, values)

            # Apply verbose style.
            style = self.get_style(self.STYLE_VERBOSE)
            self.apply_style_to_cols(style, sheet, min_col=verbose_min_col, max_col=verbose_max_col, min_row=min_row)